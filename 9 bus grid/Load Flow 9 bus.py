#connect to PowerFactory
import powerfactory as pf
import random
random.seed(1)

#Library for connecting python and excel API
import openpyxl as xl
from openpyxl.chart import LineChart, Reference

app = pf.GetApplication()
app.ClearOutputWindow()

#get active project
prj = app.GetActiveProject()
filename = prj.GetAttribute("loc_name")

slack_bus = [0]
pv_bus = [1,2]
pq_bus = [3,4,5,6,7,8]

P_list = [125,90,100,50,50,50]
Q_list = [50,30,35,10,10,10]

#get all loads
loads = app.GetCalcRelevantObjects('*.ElmLod')
load_dict = {}
for i,load in enumerate(loads):
  load_dict[load.loc_name] = load
  load_dict[load.loc_name].SetAttribute("plini", P_list[i])
  load_dict[load.loc_name].SetAttribute("qlini", Q_list[i])

#get all buses
buses = app.GetCalcRelevantObjects('*.ElmTerm')
bus_dict = {}
for bus in buses:
  bus_dict[bus.loc_name] = bus

#retrieve load-flow object
ldf = app.GetFromStudyCase("ComLdf")

wb = xl.Workbook()
sheet = wb.active
sheet.title = 'PowerFlowData'

sheet['A1'] = 'PF Dataset'

for n,bus_key in enumerate(bus_dict.keys()):
  
  if n in slack_bus:
    sheet.cell(row = 1, column = n*4+2).value = ("P_%s (slack)" % (n+1))
    sheet.cell(row = 1, column = n*4+3).value = ("Q_%s (slack)" % (n+1))
    sheet.cell(row = 1, column = n*4+4).value = ("V_%s (slack)" % (n+1))
    sheet.cell(row = 1, column = n*4+5).value = ("d_%s (slack)" % (n+1))
  
  elif n in pv_bus:
    sheet.cell(row = 1, column = n*4+2).value = ("P_%s (PV)" % (n+1))
    sheet.cell(row = 1, column = n*4+3).value = ("Q_%s (PV)" % (n+1))
    sheet.cell(row = 1, column = n*4+4).value = ("V_%s (PV)" % (n+1))
    sheet.cell(row = 1, column = n*4+5).value = ("d_%s (PV)" % (n+1))
  
  elif n in pq_bus:
    sheet.cell(row = 1, column = n*4+2).value = ("P_%s (PQ)" % (n+1))
    sheet.cell(row = 1, column = n*4+3).value = ("Q_%s (PQ)" % (n+1))
    sheet.cell(row = 1, column = n*4+4).value = ("V_%s (PQ)" % (n+1))
    sheet.cell(row = 1, column = n*4+5).value = ("d_%s (PQ)" % (n+1))
    
for i in range(10): #make 10 data point/10 load flow calculation
  app.PrintPlain("ITERATION %s" % (i+1))
  
  sheet.cell(row = i+2, column = 1).value = 'Data ' + str(i+1)
  
  for n,load_key in enumerate(load_dict.keys()):
    load_dict[load_key].SetAttribute("plini", P_list[n] + P_list[n]*random.uniform(-0.5, 0.5))
    load_dict[load_key].SetAttribute("qlini", Q_list[n] + Q_list[n]*random.uniform(-0.5, 0.5))
  
  #force balanced load flow  
  ldf.iopt_net = 0
  #execute load flow
  ldf.Execute()
  
  #for n, bus in enumerate(buses):
  for n,bus_key in enumerate(bus_dict.keys()):
    
    P_gen = bus_dict[bus_key].GetAttribute("m:Pgen")
    Q_gen = bus_dict[bus_key].GetAttribute("m:Qgen")
    V_gen = bus_dict[bus_key].GetAttribute("m:U")
    d_gen = bus_dict[bus_key].GetAttribute("m:phiu")
    P_load = bus_dict[bus_key].GetAttribute("m:Pload")
    Q_load = bus_dict[bus_key].GetAttribute("m:Qload")
    V_load = bus_dict[bus_key].GetAttribute("m:U")
    d_load = bus_dict[bus_key].GetAttribute("m:phiu")
    
    if n in slack_bus:
      sheet.cell(row = i+2, column = n*4+2).value = '{0:.3f}'.format(P_gen)
      sheet.cell(row = i+2, column = n*4+3).value = '{0:.3f}'.format(Q_gen)
      sheet.cell(row = i+2, column = n*4+4).value = '{0:.3f}'.format(V_gen)
      sheet.cell(row = i+2, column = n*4+5).value = '{0:.3f}'.format(d_gen)
    
    elif n in pv_bus:
      sheet.cell(row = i+2, column = n*4+2).value = '{0:.3f}'.format(P_gen)
      sheet.cell(row = i+2, column = n*4+3).value = '{0:.3f}'.format(Q_gen)
      sheet.cell(row = i+2, column = n*4+4).value = '{0:.3f}'.format(V_gen)
      sheet.cell(row = i+2, column = n*4+5).value = '{0:.3f}'.format(d_gen)
      
    elif n in pq_bus:
      sheet.cell(row = i+2, column = n*4+2).value = '{0:.3f}'.format(P_load)
      sheet.cell(row = i+2, column = n*4+3).value = '{0:.3f}'.format(Q_load)
      sheet.cell(row = i+2, column = n*4+4).value = '{0:.3f}'.format(V_load)
      sheet.cell(row = i+2, column = n*4+5).value = '{0:.3f}'.format(d_load)
      
#save the excel file
wb.save('C:\\Users\\ASUS\\Desktop\\%s.xlsx' % (filename))