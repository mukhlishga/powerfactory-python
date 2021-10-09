#This python code is for every kind of power grid in Power project

#connect to PowerFactory
import powerfactory as pf

app = pf.GetApplication()
app.ClearOutputWindow()

#get active project
prj = app.GetActiveProject()
filename = prj.GetAttribute("loc_name")

#get all relevant parameter
buses = app.GetCalcRelevantObjects('*.ElmTerm')
 
#retrieve load-flow object
ldf = app.GetFromStudyCase("ComLdf")
#force balanced load flow
ldf.iopt_net = 0

#execute load flow
ldf.Execute()

#Library for connecting python and excel API
import openpyxl as xl
from openpyxl.chart import LineChart, Reference
  
wb = xl.Workbook()
sheet = wb.active
sheet.title = 'PowerFlowData'

sheet['A1'] = 'Bus'
sheet['B1'] = 'Name'
sheet['C1'] = 'V (p.u.)'
sheet['D1'] = 'Angle (deg)'
sheet['E1'] = 'P (MW)'
sheet['F1'] = 'Q (MVar)'
  
i = 0

for bus in buses:
  i += 1
  app.PrintPlain("bus %s" % (bus))
  Name = bus.GetAttribute("loc_name")
  V = bus.GetAttribute("m:u")
  Angle = bus.GetAttribute("m:phiu")
  P = bus.GetAttribute("m:Pflow")
  Q = bus.GetAttribute("m:Qflow")
  #print to terminal
  app.PrintPlain("V at terminal %s is %f p.u." % (bus, V))
  app.PrintPlain("Angle at terminal %s is %f p.u." % (bus, Angle))
  app.PrintPlain("Pgen at terminal %s is %f p.u." % (bus, P))
  app.PrintPlain("Qgen at terminal %s is %f p.u." % (bus, Q))
  #write to excel file
  sheet.cell(row = 1+i, column = 1).value = i
  sheet.cell(row = 1+i, column = 2).value = Name
  sheet.cell(row = 1+i, column = 3).value = V
  sheet.cell(row = 1+i, column = 4).value = Angle
  sheet.cell(row = 1+i, column = 5).value = P
  sheet.cell(row = 1+i, column = 6).value = Q
    
#save the excel file
wb.save('C:\\Users\\ASUS\\Desktop\\%s.xlsx' % (filename))